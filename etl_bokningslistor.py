#!/usr/bin/env python3
"""
Golf Joy Travel — ETL för bokningslistor  ->  avidentifierad ban-statistik.
EXTRACT: läs alla 'Bokningslista*.xlsx' i en mapp (per-lands-blad).
TRANSFORM: hoppa ALLTID över namn/adress/mail/tel (avidentifiera). Tvätta banetiketter.
           Räkna banpopularitet och ban-PAR (co-occurrence) per destination.
LOAD: skriv SQL (bana_bokad + bana_par_statistik) för Supabase.
Kör:  python3 etl_bokningslistor.py <mapp med listor> <ut.sql>
"""
import sys, os, glob, re
from itertools import combinations
from collections import Counter
import openpyxl

PII={"namn","adress","mailadress","maildress","mail-adress","telefonnummer","telefon-nummer","telefon","mobilnummer"}
META={"idag","datum för bokning","datum bokat","avresa","avresda","hemresa","pris","hotell","boende",
"typ","info","hcp","flyg","färja","hyrbil","transfer","transfer/hyrbil","hyrbil/transfer","ai/hb",
"klubbor","utflykter dubai","toatalt anta resenärer","","flyg ","transfer/hyrbil ",
"övriga banor","övriga murcia banor","övriga banor edinburg","övriga banor "}
SKIP_SHEETS={"presentkort","kortbetalningar","valutaberäkning","övriga"}

def norm(x): return ("" if x is None else str(x)).strip().lower()

CRUFT=[r"förskottsbetalas.*$", r"förskott.*$", r"via faktura.*$", r"via kort.*$",
       r"vi bokar alltid direkt.*$", r"betalas.*$", r"\(.*?\)", r"\d+ *månad.*$"]
def clean_course(name):
    s=str(name).strip()
    for pat in CRUFT: s=re.sub(pat,"",s,flags=re.IGNORECASE)
    return re.sub(r"\s+"," ",s).strip(" -,")

def harvest_sheet(ws, pop, pairs):
    rows=[list(r) for i,r in enumerate(ws.iter_rows(values_only=True)) if i<600]
    hi=None
    for i,r in enumerate(rows):
        s={norm(c) for c in r}
        if "namn" in s and ("avresa" in s or "avresda" in s): hi=i;break
    if hi is None: return 0
    header=[norm(c) for c in rows[hi]]
    course_idx=[j for j,h in enumerate(header) if h not in META and h not in PII]
    names={j:clean_course(rows[hi][j]) for j in course_idx}
    av = header.index("avresa") if "avresa" in header else (header.index("avresda") if "avresda" in header else 1)
    nb=0
    for r in rows[hi+1:]:
        if not (av<len(r) and r[av] not in (None,"")): continue
        nb+=1
        booked=sorted({names[j] for j in course_idx if j<len(r) and r[j] not in (None,"",0) and names[j]})
        for b in booked: pop[(ws.title,b)]+=1
        for a,b in combinations(booked,2): pairs[(ws.title,a,b)]+=1
    return nb

def main(folder, out_sql, min_par=2, min_bana=2):
    files=sorted(glob.glob(os.path.join(folder,"Bokningslista*.xlsx")))
    pop=Counter(); pairs=Counter(); nfiles=0; nb_total=0
    for f in files:
        try: wb=openpyxl.load_workbook(f, read_only=True, data_only=True)
        except Exception as e: print("  hoppar",os.path.basename(f),e); continue
        nfiles+=1
        for ws in wb.worksheets:
            if norm(ws.title) in SKIP_SHEETS or ws.max_row<3: continue
            try: nb_total+=harvest_sheet(ws,pop,pairs)
            except Exception: pass
        wb.close()
    def esc(s): return "'"+str(s).replace("'","''")+"'"
    L=["-- Golf Joy — avidentifierad ban-statistik (ETL-utdata)",
       f"-- Källor: {nfiles} bokningslista(or), {nb_total} bokningar. INGA personuppgifter.","","begin;","",
       "create table if not exists bana_bokad (destination text, bana text, antal int, primary key(destination,bana));",
       "create table if not exists bana_par_statistik (destination text, bana_a text, bana_b text, antal int, primary key(destination,bana_a,bana_b));",""]
    for (dest,bana),n in sorted(pop.items(), key=lambda x:(-x[1])):
        if n<min_bana or not bana: continue
        L.append(f"insert into bana_bokad values ({esc(dest)},{esc(bana)},{n}) on conflict (destination,bana) do update set antal=excluded.antal;")
    L.append("")
    for (dest,a,b),n in sorted(pairs.items(), key=lambda x:(-x[1])):
        if n<min_par: continue
        L.append(f"insert into bana_par_statistik values ({esc(dest)},{esc(a)},{esc(b)},{n}) on conflict (destination,bana_a,bana_b) do update set antal=excluded.antal;")
    L+=["","commit;",""]
    open(out_sql,"w",encoding="utf-8").write("\n".join(L))
    print(f"Filer: {nfiles} | Bokningar: {nb_total} | Banor: {len({k[1] for k in pop})} | Par: {sum(1 for _,n in pairs.items() if n>=min_par)}")
    print("SQL:",out_sql)

if __name__=="__main__":
    folder=sys.argv[1] if len(sys.argv)>1 else "/mnt/project"
    out=sys.argv[2] if len(sys.argv)>2 else "golfjoy_banpar_relationer.sql"
    main(folder,out)
